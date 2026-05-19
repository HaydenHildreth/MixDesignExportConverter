"""
mpaq_convert_mixes.py
author  Hayden Hildreth
version 0.1.0
revision date 05/19/2026

Convert MPAQ's mix design CSV export into the flat, row-per-ingredient
format that can be imported into Keystone.

MPAQ stores every ingredient as a numeric ID (Agg1ID, Cem1ID, Adm1ID, …)
so you must supply a lookup file that maps those IDs to ingredient names.
A sample lookup CSV is shown in the Usage section below.

Output columns:
  Column A (col 0): Mix Design Name  (MixId value, uppercased)
  Column B (col 1): Ingredient name  (from lookup, or auto-generated)
  Column C (col 2): Unit             (LB for agg/cem/water, OZ for admixtures)
  Column D (col 3): Amount           (formatted to 3 decimal places)

Usage:
  python mpaq_convert_mixes.py [input.csv] [output.xlsx] [lookup.csv] [plant_separator]

Arguments:
  input.csv         – MPAQ mix design export  (default: mpaq_export.csv)
  output.xlsx       – Desired output file      (default: mpaq_converted.xlsx)
  lookup.csv        – Ingredient ID lookup     (default: mpaq_lookup.csv)
  plant_separator   – Optional suffix appended to every mix/ingredient name
                      (default: "")

Lookup file format (CSV, no header required, columns = ID, Name):
  1,SAND 2
  2,1 STONE
  3,PEASTONE
  ...

  Separate lookups per material type are supported via an optional third
  column containing the type abbreviation (AGG, CEM, ADM).  If the third
  column is absent the same table is used for all types:
  1,AGG,SAND 2
  1,CEM,HEIDELBERG
  1,ADM,AIRALON

  If no lookup file exists (or a particular ID is not found) the script
  falls back to auto-generated names like AGG_1, CEM_2, ADM_3 so the
  output is still usable.

Water:
  WaterTarget is always written as a WATER / LB row when the value is
  non-zero.  MPAQ stores water separately from the numbered ingredient
  slots so no lookup entry is needed.

Units:
  Aggregates  → LB
  Cementitious → LB
  Water        → LB
  Admixtures   → OZ  (industry convention; override via lookup if needed)
"""

import sys
import os
import csv
import pandas as pd
import openpyxl

# ---------------------------------------------------------------------------
# Command-line defaults
# ---------------------------------------------------------------------------
INPUT           = sys.argv[1] if len(sys.argv) > 1 else "mpaq_export.csv"
OUTPUT          = sys.argv[2] if len(sys.argv) > 2 else "mpaq_converted.xlsx"
LOOKUP_FILE     = sys.argv[3] if len(sys.argv) > 3 else "mpaq_lookup.csv"
PLANT_SEPARATOR = sys.argv[4] if len(sys.argv) > 4 else ""

# ---------------------------------------------------------------------------
# Material slot definitions
# ---------------------------------------------------------------------------
AGG_SLOTS = 6   # Agg1 … Agg6
CEM_SLOTS = 3   # Cem1 … Cem3
ADM_SLOTS = 8   # Adm1 … Adm8

# Default units per material type
UNIT_AGG   = "LB"
UNIT_CEM   = "LB"
UNIT_WATER = "LB"
UNIT_ADM   = "OZ"


# ---------------------------------------------------------------------------
# Lookup table loader
# ---------------------------------------------------------------------------
def load_lookup(path):
    """
    Return a dict keyed by (type, id_str) -> ingredient_name.
    type is one of 'AGG', 'CEM', 'ADM', or '*' (wildcard / any type).

    Accepted CSV formats (auto-detected by column count):
      2-col:  id, name          → stored under key ('*', id)
      3-col:  id, type, name    → stored under key (type.upper(), id)
              OR
              type, id, name    → same, order detected by whether col0 is alpha
    """
    lookup = {}
    if not os.path.isfile(path):
        print(f"[WARN] Lookup file '{path}' not found – using auto-generated names.")
        return lookup

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for lineno, row in enumerate(reader, 1):
            row = [c.strip() for c in row]
            if not row or all(c == "" for c in row):
                continue
            if len(row) == 2:
                id_str, name = row[0], row[1]
                lookup[("*", id_str)] = name
            elif len(row) >= 3:
                # Detect column order: if first token is alphabetic it's the type
                if row[0].isalpha():
                    typ, id_str, name = row[0].upper(), row[1], row[2]
                else:
                    id_str, typ, name = row[0], row[1].upper(), row[2]
                lookup[(typ, id_str)] = name
            else:
                print(f"[WARN] Lookup line {lineno} skipped (unexpected format): {row}")

    print(f"Loaded {len(lookup)} lookup entries from '{path}'.")
    return lookup


def resolve_name(lookup, mat_type, raw_id):
    """
    Look up (mat_type, id) then fall back to ('*', id) then to auto-name.
    raw_id is cast to string; float IDs like 1.0 are normalised to '1'.
    """
    # Normalise float IDs (e.g. 1.0 -> '1')
    try:
        id_str = str(int(float(raw_id)))
    except (ValueError, TypeError):
        id_str = str(raw_id).strip()

    name = (
        lookup.get((mat_type, id_str))
        or lookup.get(("*", id_str))
        or f"{mat_type}_{id_str}"
    )
    return name


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------
def parse_mixes(path, lookup):
    """
    Read the MPAQ CSV and return a list of mix dicts:
      { 'name': str, 'ingredients': [(name, amount_str, unit), ...] }
    """
    df = pd.read_csv(path, dtype=str)

    # Coerce numeric columns
    numeric_cols = (
        ["WaterTarget"]
        + [f"Agg{n}ID"     for n in range(1, AGG_SLOTS + 1)]
        + [f"Agg{n}Target" for n in range(1, AGG_SLOTS + 1)]
        + [f"Cem{n}ID"     for n in range(1, CEM_SLOTS + 1)]
        + [f"Cem{n}Target" for n in range(1, CEM_SLOTS + 1)]
        + [f"Adm{n}Target" for n in range(1, ADM_SLOTS + 1)]
    )
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    mixes = []

    for _, row in df.iterrows():
        mix_id = str(row.get("MixId", "")).strip()

        # Skip blank or header-like rows
        if not mix_id or mix_id == "0" or mix_id.lower() == "mixid":
            continue

        ingredients = []

        # -- Aggregates -------------------------------------------------------
        for n in range(1, AGG_SLOTS + 1):
            agg_id  = row.get(f"Agg{n}ID")
            agg_amt = row.get(f"Agg{n}Target")
            if pd.isna(agg_id) or pd.isna(agg_amt):
                continue
            try:
                amt_f = float(agg_amt)
            except (ValueError, TypeError):
                continue
            if amt_f == 0:
                continue
            name = resolve_name(lookup, "AGG", agg_id)
            ingredients.append((name, f"{amt_f:.3f}", UNIT_AGG))

        # -- Cementitious -----------------------------------------------------
        for n in range(1, CEM_SLOTS + 1):
            cem_id  = row.get(f"Cem{n}ID")
            cem_amt = row.get(f"Cem{n}Target")
            if pd.isna(cem_id) or pd.isna(cem_amt):
                continue
            try:
                amt_f = float(cem_amt)
            except (ValueError, TypeError):
                continue
            if amt_f == 0:
                continue
            name = resolve_name(lookup, "CEM", cem_id)
            ingredients.append((name, f"{amt_f:.3f}", UNIT_CEM))

        # -- Admixtures -------------------------------------------------------
        for n in range(1, ADM_SLOTS + 1):
            adm_id  = row.get(f"Adm{n}ID")
            adm_amt = row.get(f"Adm{n}Target")
            if pd.isna(adm_id) or adm_id == "":
                continue
            if pd.isna(adm_amt):
                continue
            try:
                amt_f = float(adm_amt)
            except (ValueError, TypeError):
                continue
            if amt_f == 0:
                continue
            name = resolve_name(lookup, "ADM", adm_id)
            ingredients.append((name, f"{amt_f:.3f}", UNIT_ADM))

        # -- Water ------------------------------------------------------------
        water_amt = row.get("WaterTarget")
        try:
            water_f = float(water_amt)
        except (ValueError, TypeError):
            water_f = 0.0
        if water_f != 0:
            ingredients.append(("WATER", f"{water_f:.3f}", UNIT_WATER))

        if ingredients:
            mixes.append({"name": mix_id, "ingredients": ingredients})

    return mixes


# ---------------------------------------------------------------------------
# Writer
# ---------------------------------------------------------------------------
def write_output(mixes, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    out_row = 1  # openpyxl is 1-indexed

    for mix in mixes:
        name = (mix["name"] + PLANT_SEPARATOR).upper()
        for (ingredient, amount, unit) in mix["ingredients"]:
            ws.cell(out_row, 1, name)
            ws.cell(out_row, 2, (ingredient + PLANT_SEPARATOR).upper())
            ws.cell(out_row, 3, unit.upper())
            ws.cell(out_row, 4, amount)
            out_row += 1

    wb.save(path)
    print(f"Written {out_row - 1} rows across {len(mixes)} mixes -> {path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print(f"Parsing '{INPUT}' ...")
    lookup = load_lookup(LOOKUP_FILE)
    mixes  = parse_mixes(INPUT, lookup)
    print(f"Found {len(mixes)} mix designs.")
    if PLANT_SEPARATOR:
        print(f"Using plant separator: {repr(PLANT_SEPARATOR)}")
    write_output(mixes, OUTPUT)
