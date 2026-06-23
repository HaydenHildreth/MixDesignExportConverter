"""
sysdyne_convert_mixes.py
author Hayden Hildreth
version 1.2.0
revision date 06/23/2026

Convert sysdyne's mix design XLSX export into a format importable into Keystone.

Input format (XLSX, header on row 3, data from row 4):
  Each mix+plant combination begins with a row where 'Item Code' is populated.
  Subsequent ingredient rows have NaN in 'Item Code' and inherit the current mix context.
  Columns of interest:
    Item Code                     - Mix design code
    Description                   - Mix design description (col B in output)
    Location                      - Plant number, appended to mix code in col A
    Constituent Short Description - Ingredient name
    Quantity                      - Amount
    Unit of Measure               - Unit (ga -> GL, others uppercased)

  The input file also contains a secondary sheet ('Sheet1') with a list of mix codes
  (in column A, rows 14-211) that are highlighted yellow via a conditional formatting
  rule. This script reads that list and re-applies yellow highlighting to any matching
  rows in the output.

Output format (6 columns, matching Keystone import spec):
  Column A (col 0): Mix code + product_separator + plant number + plant_separator
  Column B (col 1): Mix Description
  Column C (col 2): Ingredient name
  Column D (col 3): (empty)
  Column E (col 4): Amount  (2 decimal places)
  Column F (col 5): Unit (LB / OZ / GL)

  Output is written as .xlsx so cell fill colors are supported.

Usage:
  python newvendor_convert_mixes.py [input.xlsx] [output.xlsx] [plant_separator] [product_separator]

Defaults (used if script is run without parameters at runtime):
  input             = original_format.xlsx
  output            = output_converted.xlsx
  plant_separator   = ""
  product_separator = ""

Arguments:
  plant_separator   - String appended to the very end of the mix code in col A.
                      Matches the convention used in command_convert_mixes.py and
                      mpaq_convert_mixes.py.
  product_separator - String inserted between the Item Code and the plant number
                      from the Location column.
                      Example: product_separator="-" -> "100SLHP-01"
                      If omitted, the plant number is still appended with no
                      separator: "100SLHP01"

Notes:
  - Ingredients with a zero or blank quantity are skipped.
  - Unit 'ga' is normalized to 'GL' to match Keystone convention; all units uppercased.
  - Each unique Item Code + Location combination becomes its own group in the output.
  - Rows whose Item Code appears in the highlight list (Sheet1 col A) are written
    with a yellow background, matching the original file's conditional formatting.
"""

import sys
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import warnings

warnings.filterwarnings("ignore", category=UserWarning)

INPUT             = sys.argv[1] if len(sys.argv) > 1 else "original_format.xlsx"
OUTPUT            = sys.argv[2] if len(sys.argv) > 2 else "output_converted.xlsx"
PLANT_SEPARATOR   = sys.argv[3] if len(sys.argv) > 3 else ""
PRODUCT_SEPARATOR = sys.argv[4] if len(sys.argv) > 4 else ""

# Row index of the header in the XLSX (0-indexed; row 3 in Excel = index 2)
HEADER_ROW = 2

# Sheet and range where highlighted mix codes are stored
HIGHLIGHT_SHEET    = "Sheet1"
HIGHLIGHT_COL      = 1        # Column A 1-indexed
HIGHLIGHT_ROW_MIN  = 14
HIGHLIGHT_ROW_MAX  = 211

YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

UNIT_MAP = {
    "ga": "GL",  # gallons
}


def normalize_unit(unit_str):
    """Uppercase and normalize unit strings to Keystone conventions."""
    u = str(unit_str).strip()
    return UNIT_MAP.get(u.lower(), u.upper())


def load_highlighted_codes(path):
    """
    Read the list of mix codes from Sheet1 col A (rows 14-211) that should
    be highlighted yellow in the output.
    Returns a set of uppercased code strings.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if HIGHLIGHT_SHEET not in wb.sheetnames:
        print(f"Warning: '{HIGHLIGHT_SHEET}' not found — no highlighting will be applied.")
        return set()

    ws = wb[HIGHLIGHT_SHEET]
    codes = set()
    for row in ws.iter_rows(
        min_row=HIGHLIGHT_ROW_MIN, max_row=HIGHLIGHT_ROW_MAX,
        min_col=HIGHLIGHT_COL,    max_col=HIGHLIGHT_COL,
        values_only=True
    ):
        val = row[0]
        if val and str(val).strip():
            codes.add(str(val).strip().upper())

    wb.close()
    return codes


def parse_mixes(path):
    """
    Read the vendor XLSX into a list of mix dicts:
      {
        'code': str,   - Item Code
        'plant': str,  - Location (plant number)
        'desc': str,   - Description
        'ingredients': [(name, amount_str, unit), ...]
      }

    The file uses a fill-down pattern: only the first ingredient row of each
    mix+plant group has Item Code / Description / Location populated. Subsequent
    rows within that group have NaN in those columns and inherit the current context.

    A new mix+plant group begins whenever Item Code is non-empty.
    Ingredients with quantity == 0 are skipped.
    """
    df = pd.read_excel(path, header=HEADER_ROW, dtype=str, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]

    mixes = []
    current_code  = None
    current_plant = None
    current_desc  = None
    current_ingredients = []

    def flush():
        if current_code and current_ingredients:
            mixes.append({
                "code":        current_code,
                "plant":       current_plant,
                "desc":        current_desc,
                "ingredients": list(current_ingredients),
            })

    for _, row in df.iterrows():
        item_code = str(row.get("Item Code", "")).strip()
        is_new_group = item_code != "" and item_code != "nan"

        if is_new_group:
            flush()
            current_code  = item_code
            current_plant = str(row.get("Location", "")).strip()
            current_desc  = str(row.get("Description", "")).strip()
            current_ingredients = []

        if current_code is None:
            continue

        ingredient = str(row.get("Constituent Short Description", "")).strip()
        if not ingredient or ingredient == "nan":
            continue

        qty_raw = row.get("Quantity", None)
        try:
            qty_f = float(qty_raw) if not pd.isna(qty_raw) else 0.0
        except (ValueError, TypeError):
            qty_f = 0.0

        if qty_f == 0.0:
            continue

        unit = normalize_unit(row.get("Unit of Measure", ""))
        current_ingredients.append((ingredient, f"{qty_f:.2f}", unit))

    flush()
    return mixes


def build_mix_code(code, plant):
    """
    Assemble the full col A value:
      {ItemCode}{PRODUCT_SEPARATOR}{Plant}{PLANT_SEPARATOR}
    All uppercased.
    """
    result = code.upper()
    if plant and plant != "nan":
        result += PRODUCT_SEPARATOR + plant.upper()
    result += PLANT_SEPARATOR
    return result


def write_output(mixes, path, highlighted_codes):
    """
    Write mixes to a .xlsx file matching the Keystone import format:
      Col A: Mix code  |  Col B: Description  |  Col C: Ingredient  |
      Col D: (empty)   |  Col E: Amount        |  Col F: Unit

    Rows whose base Item Code is in highlighted_codes get a yellow background.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    out_row = 1  # 1-index
    highlighted_count = 0

    for mix in mixes:
        full_code  = build_mix_code(mix["code"], mix["plant"])
        desc       = mix["desc"].upper()
        # Check the base Item Code (without separators) against the highlight list
        apply_fill = mix["code"].upper() in highlighted_codes

        for (ingredient, amount, unit) in mix["ingredients"]:
            ws.cell(out_row, 1, full_code)
            ws.cell(out_row, 2, desc)
            ws.cell(out_row, 3, ingredient.upper())
            # Column D intentionally left empty (col index 4)
            ws.cell(out_row, 5, float(amount))
            ws.cell(out_row, 6, unit)

            if apply_fill:
                for col in range(1, 7):
                    ws.cell(out_row, col).fill = YELLOW_FILL
                highlighted_count += 1

            out_row += 1

    wb.save(path)
    total_rows = out_row - 1
    print(f"Written {total_rows} rows across {len(mixes)} mixes -> {path}")
    print(f"Highlighted {highlighted_count} rows yellow ({len(highlighted_codes)} codes in highlight list)")


if __name__ == "__main__":
    print(f"Parsing {INPUT} ...")
    highlighted_codes = load_highlighted_codes(INPUT)
    print(f"Loaded {len(highlighted_codes)} highlighted mix codes from '{HIGHLIGHT_SHEET}'.")
    mixes = parse_mixes(INPUT)
    print(f"Found {len(mixes)} mix designs.")
    if PRODUCT_SEPARATOR:
        print(f"Using product separator: {repr(PRODUCT_SEPARATOR)}")
    if PLANT_SEPARATOR:
        print(f"Using plant separator:   {repr(PLANT_SEPARATOR)}")
    write_output(mixes, OUTPUT, highlighted_codes)
